from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver import ActionChains
from time import sleep
import pandas as pd
import glob
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter ,column_index_from_string

import numpy as np
import os

import pickle
import imaplib
import email
from email.header import decode_header



f = open('test.txt')
pol = 0
for line in f:
    if pol == 0:

        datestart = line.rstrip()
        print("время запуска отчёта", datestart)
    if pol == 1:
        dateend = line.rstrip()
        print("время конца отчёта", dateend)
    pol = pol + 1
    if pol > 2:
#password:usernamed:id:nazvanie 
        passw = line[:line.find(":")]  # этак конструкция печатает line(строку) заканчивая элементом : в этой строке
        passw1 = line[:line.find(":")+1]
        ostov = line.replace(passw1, "")
        user = ostov[:ostov.find(":")]
        user1 = ostov[:ostov.find(":")+1]
        ostov2 = ostov.replace(user1, "")
        ID = ostov2[:ostov2.find(":")]
        nazvan = ostov2.replace(ID, "")
        nazvanie = nazvan[nazvan.find(":")+1:] # имя фирмы
        #print(passw, "пароль")
        #rint(ostov, "")
        #print(user, "Имя пользователя")
        #print(ID)ааа
        #print(nazvanie)
       
        username = "***********@gmail.com"
        password = "******"

        imap = imaplib.IMAP4_SSL("imap.gmail.com")
        # authenticate
        imap.login(username, password)
        # select the mailbox I want to delete in
        # if you want SPAM, use imap.select("SPAM") instead
        imap.select("INBOX")
        status, messages = imap.search(None, "ALL")
        messages = messages[0].split(b' ')



        provekafull = 0
        while provekafull == 0:
            try:


        
                #breakpoint()
                options = Options()

                options.add_argument("--disable-notifications");
                driver = webdriver.Chrome(executable_path=r"chromedriver.exe", options=options)

                #options.add_argument("download.default_directory=F:\Scripts")
                #options = Options()
                #options.add_argument("download.default_directory=F:\Scripts")
                #options.add_argument("--disable-notifications");
                #options.add_argument("download.default_directory=F:\Scripts")
                #options.add_argument('--headless')
                #options.add_argument('--disable-gpu')  # Last I checked this was necessary.
                #System.setProperty("webdriver.chrome.driver","F:\Scripts\chromedriver.exe");

                check3 = 11
               # while check3 < 30:   #while try except позволит пытаться пройти цикл бесконечное количество раз. Т.к. код в 70-80% проходит, то нормально всё.
                    #try:
                driver.implicitly_wait(10)

                sleeptime = 0
                while sleeptime < 60:
                    try:
                        
                        driver.get("https://www.1-ofd.ru")
                        sleeptime = 60
                    except:
                        sleeptime = sleeptime + 1
                        time.sleep(60)
                        driver.refresh()
                        print("перезагружаю страницу https://www.1-ofd.ru")

                
                
                #assert "Python" in driver.title
                         #   try:с
                          #      timer1 = str(proc.create_time())
                           # except psutil.NoSuchProcess:
                           #     print("ошипка2")
                time.sleep(1)


                print(nazvanie)
                sleeptime = 0
                while sleeptime < 60:
                    try:
                        
                        driver.find_element_by_xpath("/html/body/div/div/div[1]/header/div/div[2]/button[2]").click()
                        sleeptime = 60
                    except:
                        sleeptime = sleeptime + 1
                        time.sleep(15)
                        driver.refresh()
                        
                        
                try:
                    username = driver.find_element_by_name("login")
                except:
                    time.sleep(3)
                    driver.find_element_by_xpath("/html/body/div/div/div[1]/header/div/div[2]/button[2]").click()
                    username = driver.find_element_by_name("login")
                #впиндюрить цикл сдеся
                username.send_keys(user)
                #/html/body/div[4]/div/div/div[2]/form/div[3]/button
                driver.implicitly_wait(1)


                #https://chrome.google.com/webstore/detail/xpath-finder/ihnknokegkbpmofmafnkoadfjkhlogph/related

                try:

                    driver.find_element_by_xpath("/html/body/div[6]/div/div/div[2]/form/div[3]/button").click()
                except :
                    driver.find_element_by_xpath("/html/body/div[7]/div/div/div[2]/form/div[3]/button").click()
                    
                psswd = driver.find_element_by_xpath("//*[@id='form-authorization-password']")
                psswd.send_keys(passw)
                try:
                    driver.find_element_by_xpath("/html/body/div[6]/div/div/div[2]/form/div[4]/button").click()
                except :
                    driver.find_element_by_xpath("/html/body/div[7]/div/div/div[2]/form/div[4]/button").click()
                check = 0
                check43 = 0
                while check43 < 10:
                    try:
                        
                
                        element = driver.find_element_by_xpath("//*[contains(text(), 'Мониторинг')]")
                        check43 = 11
                    except:
                        sleep(30)
                        driver.refresh()
                        print("Ищю слово Мониторинг на сайте ОФД, это проверка, что он загрузился")
                        check43 = check43 + 1
                        
                while check == 0:
                    try:
                        sleep(5)
                        check = driver.find_element_by_xpath("/html/body/app-root/div/div/div[2]/app-coach-onboarding/div/div[2]/div/app-onboard-item/div/div[1]/button").click() #кнопка пропустить когда спрашивают верный ли эмаил
                        check = 1
                    except:
                        
                        try:
                            driver.find_element_by_xpath("/html/body/app-root/div/div/div[2]/app-coach-onboarding/div/div[2]/div/app-onboard-item/div/div[1]/buttons").click() #кнопка пропустить когда спрашивают верный ли эмаил
                            check = 1
                        except:
                            time.sleep(15)
                            check = 1
                            print("не вижу кнокки пропустить верны ли эмаил")
                            driver.refresh()
                            
                print(element)
                element_X = driver.find_elements_by_xpath("//body")
                for x in element_X[2:]:  # Эта конструкция начнёт поиск с 3 го элемента списка и до конца
                    x.click()

                try:
                    driver.find_element_by_xpath("//body/div/div/div/button")
                    print("ликну,чтобы скрыть бота помошника")
                except:
                    sleep(1)
                  
                

                #driver.find_element_by_xpath("/html/body/app-root/div/div/div[2]/app-monitoring/div/app-monitoring-kkms/div[1]/app-filters/div/div[2]/app-collapse-control/div/div[2]/div/app-filter-tags/div/div[3]/app-filter-tag[1]/app-filter-tag-simple/div/span[1]").click()

                driver.get("https://org.1-ofd.ru/reports")
                #driver.find_element_by_tag_name('app-flat-button').click()
                proverka_ofd = 0
                sleep(10)
                try:
                    driver.find_element_by_xpath("//html/body/div/div/div/button/span").click()
                    print("ликну,чтобы скрыть бота помошника")
                except:
                    sleep(1)
                print("не вижу на сайте отчётов ничего")
                driver.refresh()
                sleep(2)
                #breakpoint()  
                pafinder = '//*[@id="undefined"]'
                driver.find_element_by_xpath(pafinder).click()
                driver.find_element_by_xpath("//div/div/div[7]/div").click()
                driver.find_element_by_xpath("//div[2]/div[3]/app-flat-button").click()
                time.sleep(3)
                print("галочка выбрать все кассы")
                driver.find_element_by_xpath("//div/div[2]/div[2]/div[1]/div[1]/div/div[1]/app-custom-checkbox/div/span").click()   # галочка выбрать все кассы
                time.sleep(3)
                print("галочка подтвердить выбор")
                driver.find_element_by_xpath("//app-select-kkms-multiple/div/div[3]/div/app-action-button/div/div/div").click()# галочка подтвердить выбор
                time.sleep(3)
                #time.sleep(3)
                #driver.find_element_by_xpath("/html/body/app-root/div/div/div[2]/app-reports-page/div/div[2]/app-report-params/div/div[2]/div[3]/div/svg/").click()
                #driver.find_element_by_class_name('ng-tns-c269-2').click()
                time.sleep(1)
                email1 = driver.find_element_by_xpath("//*[@id='email']")
                time.sleep(1)
                email1.clear()
                time.sleep(1)
                print("Ввожу емаил")
                email1.send_keys("testchekget@gmail.com")
                time.sleep(1)

                #старый не рабочий ввод даты начала и конца
                #start_date = driver.find_element_by_xpath("//div[1]/app-date-time-picker/div/div/input")
                time.sleep(1)
                #start_date.click()
                time.sleep(1)
                #start_date.send_keys(datestart)
                #end_date = driver.find_element_by_xpath("//div[3]/app-date-time-picker/div/div/input")
                #end_date.click()
                #end_date.send_keys(dateend)
                nachalo = "Начало"
                konec = "Конец"
                start_date = driver.find_element_by_xpath("//*[@placeholder='" + nachalo + "']")
                start_date.send_keys(datestart)
                end_date = driver.find_element_by_xpath("//*[@placeholder='" + konec + "']")
                end_date.send_keys(dateend)
                driver.find_element_by_xpath("/html/body/app-root/div/div/div[2]/app-reports-page/div/div[2]/app-report-params/div/div[2]/div[2]/app-flat-button/div").click() #кнопка добавить когда добавляешь эмаил в рассылку

                #driver.find_element_by_xpath("/html/body/app-root/app-header/div/div").remove()#тест ремува элемента
                driver.find_element_by_xpath("/html").send_keys(u'\ue00f')
                sleep(2)
                #driver.find_element_by_xpath("//html/body/div/div[1]/div[1]/button/span").click()#кнопка закрыть уведомления чат бота

                driver.find_element_by_xpath("/html/body/app-root/div/div/div[2]/app-reports-page/div/div[2]/app-report-params/div/div[4]/app-action-button/div/div/div").click()#кнопка заказать отчёт

                #driver.find_element_by_xpath("/html/body/app-root/div/div/div[2]/app-reports-page/div/div[2]/app-report-params/div/div[2]/div[2]/app-flat-button/div").click()#кнопка заказать отчёт

                sleep(3)
                # элемент покрыт другим элементом, обход ошибки
                proverka_ofd = 1


                


                proverka_google = 1
                while proverka_google == 1:
                    try:
                        username = '**********@gmail.com'
                        password = '********'

                        #driver = webdriver.Chrome(r"F:\Scripts\chromedriver.exe")
                        #driver = webdriver.Chrome(r"F:\Scripts\chromedriver.exe")    
                        driver.get('https://stackoverflow.com/users/signup?ssrc=head&returnurl=%2fusers%2fstory%2fcurrent%27')
                        sleep(3)
                        driver.find_element_by_xpath('//*[@id="openid-buttons"]/button[1]').click()
                        driver.find_element_by_xpath('//input[@type="email"]').send_keys(username)
                        driver.find_element_by_xpath('//*[@id="identifierNext"]').click()
                        sleep(3)
                        driver.find_element_by_xpath('//input[@type="password"]').send_keys(password)
                        driver.find_element_by_xpath('//*[@id="passwordNext"]').click()
                        sleep(2)
                        driver.get('https://mail.google.com/mail/u/0/?tab=rm&ogbl#inbox')
                        sleep(5)

                        text1 = '1ОФД. Отчет по чекам с позициями '
                        year = '2021-'
                        month = '10-'
                        date = '29'
                        time1 = 0000
                        result = text1 + year + month + date
                        sleep(5)
                       ##### self.driver.find_element_by_xpath("//tbody//tr/td[5]/div/div/div[2]/span/span[contains(text(),'1ОФД. Отчет по чекам с позициями ' )]")
                       # try:
                      #      self.driver.find_element_by_xpath('/html/body/div[7]/div[3]/div/div[2]/div[1]/div[2]/div/div/div/div/div[1]/div[2]/div[1]/div[1]/div/div/div[1]/div/div[1]').click()
                       # except:
                       #     self.driver.find_element_by_xpath('/html/body/div[7]/div[3]/div/div[2]/div[1]/div[2]/div/div/div/div/div[1]/div/div[1]/div[1]/div/div/div[1]/div/div[1]/span').click()
                        # поиск упоминаний 1ОФД в строках. Двойно слеш в середине пути означает что будем искать во всех tr из tbody
                    # поиск упоминаний 1ОФД в строках. Двойной слеш в середине пути означает что будем искать во всех tr из tbody
                        check2 = 0
                        while check2 == 0:
                            
                            try:
                                driver.refresh()
                                sleep (30)
                                element = driver.find_elements_by_xpath("//*[contains(text(), '1ОФД. Отчет по чекам с позициями')]")

                                print("шо за хня кликать")
                                print(element)
                                for i in element:
                                    print("начинаю кликать")
                                    try:
                                        print(i.get_attribute("value"))
                                        i.click()
                                        print("кликнул")
                                        check2 = 1
                                    except:
                                        print("Не вижу печать очёта по чекам")
                            except:
                                sleep (10)
                                driver.refresh()

                            #except:
                             #   sleep (10)
                              #  driver.refresh()
                        try:
                            razvernut_spisok_email = driver.find_element_by_class_name('ajR')
                            razvernut_spisok_email.click()
                        except:
                            sleep(3)
                        
                        #driver.find_element_by_link_text("загрузить отчет из личного кабинета").click()
                        sleep(5)
                        

                        driver.find_element_by_link_text("загрузить отчет из личного кабинета").click()
                        sleep(5)
                        
                        #driver.find_element_by_xpath('//*[@id=":6k"]').click()


#Ниже вставка для удаления всех сообщений в почте








                        #driver.find_element_by_link_text("Удалить это письмо").click()        
                        

                        print("типо убил")
                        
                        sleep(1)
                        #breakpoint()



                        
                        proverka_google = 0





                        

                    except:
                        
                        sleep(1)
                        driver.close()







































                print("создаёт список похожих имён")
                sleep(20)
                name = glob.glob(r"C:\Users\user\Downloads\*.xlsx") # создаёт список похожих имён
                for i in name:
                    name2 = glob.escape(i)
                    file  = name2
                    print("нашёл фаил эксель", file)


                # Load spreadsheet
                #x1 = pd.ExcelFile(file)
                #print(x1)

                # Print the sheet names
                #print(xl.sheet_names)
                print("начало считывания фаила ексель")
                chekc_zagruzok = 0
                while chekc_zagruzok == 0:
                    
                    try:
                        
                        wb = load_workbook(file)
                        #print(wb.get_sheet_names())
                        sheets = wb.get_sheet_names()
                        chekc_zagruzok = 1
                    except:
                        sleep(1)
                for l in sheets:
                    sheet3 = l
                sheet = wb.get_sheet_by_name(sheet3)
                A = str(sheet.max_row)# максимальное количество строк
                B = sheet.max_column # максимальное количество колонок
                #print(sheet.max_column)
                #print(B)
                #print(B)
                B = get_column_letter(B)
                #B =scolumn_index_from_string(410)
                #print(B)
                f2 = open("reestr.txt", "a")
                kassa = [0] * 0
                problemma = [0] * 0
                organization = [0] * 0
                IDorg = [0] * 0
                print("начинаю поиск ищу кассипров")
                for cellObj in sheet['A1':B + A]:
                    for cell in cellObj:
                        
                        if cell.value == "Системный администратор":
                            colonka = cell.column

                            stroka = str(cell.row)
                            #print(colonka)
                            #print(stroka)
                            cell.value
                            nomerkassy = sheet.cell(row = cell.row, column=1).value
                            #print(sheet.cell(row = cell.row, column=1).value)
                            kassa.append(nomerkassy)
                            problemma.append(cell.value)
                            organization.append(nazvanie)
                            IDorg.append(ID)
                            print("нашёл")
                        if cell.value == "КАССИР: Администратор":
                            colonka = cell.column

                            stroka = str(cell.row)
                            #print(colonka)
                            #print(stroka)
                            cell.value
                            nomerkassy = sheet.cell(row = cell.row, column=1).value
                            #print(sheet.cell(row = cell.row, column=1).value)
                            kassa.append(nomerkassy)
                            problemma.append(cell.value)
                            organization.append(nazvanie)
                            IDorg.append(ID)
                            print("нашёл")
                        if cell.value == "Кассир 1":
                            colonka = cell.column

                            stroka = str(cell.row)
                            #print(colonka)
                            #print(stroka)
                            cell.value
                            nomerkassy = sheet.cell(row = cell.row, column=1).value
                            #print(sheet.cell(row = cell.row, column=1).value)
                            kassa.append(nomerkassy)
                            problemma.append(cell.value)
                            organization.append(nazvanie)
                            IDorg.append(ID)
                            print("нашёл")
                        if cell.value == "admin":
                            colonka = cell.column

                            stroka = str(cell.row)
                            #print(colonka)
                            #print(stroka)
                            cell.value
                            nomerkassy = sheet.cell(row = cell.row, column=1).value
                            #print(sheet.cell(row = cell.row, column=1).value)
                            kassa.append(nomerkassy)
                            problemma.append(cell.value)
                            organization.append(nazvanie)
                            IDorg.append(ID)
                            print("нашёл")
                        if cell.value == "КАССИР: Сидорова":
                            colonka = cell.column

                            stroka = str(cell.row)
                            #print(colonka)
                            #print(stroka)
                            cell.value
                            nomerkassy = sheet.cell(row = cell.row, column=1).value
                            #print(sheet.cell(row = cell.row, column=1).value)
                            kassa.append(nomerkassy)
                            problemma.append(cell.value)
                            organization.append(nazvanie)
                            IDorg.append(ID)
                            print("нашёл")
                            


                        if cell.value == "СИС. АДМИНИСТРАТОР":
                            colonka = cell.column

                            stroka = str(cell.row)
                            #print(colonka)
                            #print(stroka)
                            cell.value
                            nomerkassy = sheet.cell(row = cell.row, column=1).value
                            #print(sheet.cell(row = cell.row, column=1).value)
                            kassa.append(nomerkassy)
                            problemma.append(cell.value)
                            organization.append(nazvanie)
                            IDorg.append(ID)

                        if cell.value == "КАССИР: Иванова":
                            colonka = cell.column

                            stroka = str(cell.row)
                            #print(colonka)
                            #print(stroka)
                            cell.value
                            nomerkassy = sheet.cell(row = cell.row, column=1).value
                            #print(sheet.cell(row = cell.row, column=1).value)
                            kassa.append(nomerkassy)
                            problemma.append(cell.value)
                            organization.append(nazvanie)
                            IDorg.append(ID)
                            
                            print("нашёл")
                        if cell.value == "Продавец":
                            colonka = cell.column

                            stroka = str(cell.row)
                            #print(colonka)
                            #print(stroka)
                            cell.value
                            nomerkassy = sheet.cell(row = cell.row, column=1).value
                            #print(sheet.cell(row = cell.row, column=1).value)
                            kassa.append(nomerkassy)
                            problemma.append(cell.value)
                            organization.append(nazvanie)
                            IDorg.append(ID)
                            print("нашёл")
                        if cell.value == "КАССИР: Петрова":
                            colonka = cell.column

                            stroka = str(cell.row)
                            #print(colonka)
                            #print(stroka)
                            cell.value
                            nomerkassy = sheet.cell(row = cell.row, column=1).value
                            #print(sheet.cell(row = cell.row, column=1).value)
                            kassa.append(nomerkassy)
                            problemma.append(cell.value)
                            organization.append(nazvanie)
                            IDorg.append(ID)
                            print("нашёл")
                #print(kassa)
                            
                #print(problemma)
                massiv = np.column_stack((kassa,problemma,IDorg,organization))
                new_array = [tuple(row) for row in massiv]
                uniques = np.unique(new_array, axis=0)
                for row in uniques:
                    f2.write(str(row) +"\n")
                    print("записал")


                f2.close()
                os.remove(file)                    #stroka1 = 'A' + stroka
            #except:

                    #for clell2Obj in sheet ['A' + stroka:'AS' + stroka]:
                    #for cell2Obj in sheet['A' + stroka:'AS' + stroka]: # если находит системный  администратор, то ножно напечатать всю строку с датой чека но их будет много.
                        #for cell in cell2Obj:
                            #print("логин пароль офд")
                          #  print(cell.value, end='')# печатает в ту же строку что и предыдушее  

                
                driver.quit()


                check3 = 30
                print("зевершения цикла переход на следующее имя /пароль")
            #df1 = pd.ExcelFile(file).parse('Лист 0')
            # Load a sheet into a DataFrame by name: df1
            #df1 = xl.parse('Sheet1')


                print("удаление писем")

            
                try:
        
                    # create an IMAP4 class with SSL 
                    imap = imaplib.IMAP4_SSL("imap.gmail.com")
                    # authenticate
                    imap.login(username, password)
                    imap.select("INBOX")
                    status, messages = imap.search(None, "ALL")
                    messages = messages[0].split(b' ')
                    for mail in messages:
                        _, msg = imap.fetch(mail, "(RFC822)")
                        # вы можете удалить цикл for для повышения производительности, если у вас длинный список писем
                        # потому что он предназначен только для печати SUBJECT целевого электронного письма, которое нужно удалить
                        for response in msg:
                            if isinstance(response, tuple):
                                msg = email.message_from_bytes(response[1])
                                # расшифровать тему письма
                                subject = decode_header(msg["Subject"])[0][0]
                                if isinstance(subject, bytes):
                                    # if it's a bytes type, decode to str
                                    subject = subject.decode()
                                print("Deleting", subject)
                        # отметить письмо как удаленное
                        imap.store(mail, "+FLAGS", "\\Deleted")
                    imap.expunge()
                    # закрыть почтовый ящик
                    imap.close()
                    # выйти из аккаунта
                    imap.logout()

                except:
                    sleep(1)





                
                
                provekafull = 1


            except Exception as err:
                print(err)
                driver.quit()
                try:
                    os.remove(file)
                except:
                    sleep(1)

                try:
                    

                    # create an IMAP4 class with SSL 
                    imap = imaplib.IMAP4_SSL("imap.gmail.com")
                    # authenticate
                    imap.login(username, password)
                    imap.select("INBOX")
                    status, messages = imap.search(None, "ALL")
                    messages = messages[0].split(b' ')
                    for mail in messages:
                        _, msg = imap.fetch(mail, "(RFC822)")
                        # вы можете удалить цикл for для повышения производительности, если у вас длинный список писем
                        # потому что он предназначен только для печати SUBJECT целевого электронного письма, которое нужно удалить
                        for response in msg:
                            if isinstance(response, tuple):
                                msg = email.message_from_bytes(response[1])
                                # расшифровать тему письма
                                subject = decode_header(msg["Subject"])[0][0]
                                if isinstance(subject, bytes):
                                    # if it's a bytes type, decode to str
                                    subject = subject.decode()
                                print("Deleting", subject)
                        # отметить письмо как удаленное
                        imap.store(mail, "+FLAGS", "\\Deleted")
                    imap.expunge()
                    # закрыть почтовый ящик
                    imap.close()
                    # выйти из аккаунта
                    imap.logout()
                    name31 = glob.glob(r"C:\Users\user\Downloads\*.xlsx") # создаёт список похожих имён
                    for i in name31:
                        name4 = glob.escape(i)
                        file31  = name4
                        os.remove(file31)
                except:
                    sleep(1)
                

            #try:
            #driver.find_element_by_xpath("/html/body/div[1]/div/div/div[7]/div").click()

        #driver.find_element_by_xpath("/html/body/div[6]/div/div/div[2]/form/div[4]/button").click()
        #впиндюрить цикл сдеся
        #username.send_keys("250201001")
f.close()


